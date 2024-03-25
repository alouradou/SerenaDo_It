import os
import unittest
import openpyxl
from sqlalchemy import text, create_engine
from datetime import datetime, timedelta

from src.create_timetable import CreateTimetable
from src.data_manager import DataManager
from src.personalize_timetable import PersonalizeTimetable

os.chdir('..')


def test_db_connection(engine):
    try:
        with engine.connect():
            print("Connection to db successful")
            return True
    except Exception as e:
        print("Connection to db failed: ", e)
        return False


class TestCreateTimetable(unittest.TestCase):

    def test_fetch_google_sheet(self):
        planning_data_manager = DataManager("",
                                            sheet_name="année",
                                            saved_workbook_path="./uploads/edt Do_It.23-24.xlsx")
        choices_data_manager = DataManager("",
                                           sheet_name="effectif",
                                           saved_workbook_path="./uploads/choix.xlsx")
        planning_data_manager.load_workbook()
        choices_data_manager.load_workbook()
        self.assertTrue(os.path.exists("./uploads/choix.xlsx"))
        self.assertTrue(os.path.exists("./uploads/edt Do_It.23-24.xlsx"))

        with open("./uploads/choix.xlsx", "rb") as f:
            self.assertTrue("effectif" in openpyxl.load_workbook(f, data_only=True).sheetnames)

        with open("./uploads/edt Do_It.23-24.xlsx", "rb") as f:
            self.assertTrue("année" in openpyxl.load_workbook(f, data_only=True).sheetnames)

    def test_cache_google_sheet_db(self):
        db_engine = create_engine('sqlite:///uploads/serenadoit.db')
        self.assertTrue(test_db_connection(db_engine))

        with db_engine.connect() as conn:
            fetch_dates = conn.execute(text("SELECT DISTINCT fetch_date FROM fetch_dates ORDER BY fetch_date")).fetchall()
            last_fetch = fetch_dates[-1][0]
            last_fetch = datetime.strptime(last_fetch, "%Y-%m-%d %H:%M:%S.%f")
            print(last_fetch)

        print("Using cached data")
        with db_engine.connect() as conn:
            df = conn.execute(text("SELECT * FROM fetch_dates WHERE fetch_date=:last_fetch"),
                              {"last_fetch": last_fetch}).fetchall()


    def set_up(self):
        # Créer une instance de CreateTimetable avec des fichiers de test
        self.timetable = CreateTimetable(
            edt_file_name='./uploads/edt Do_It.23-24.xlsx',
            edt_sheet_name='année',
            students_file_name='./static/header_etudiants_do_it.xlsx',
            students_sheet_name='effectif'
        )
        # Initialiser la feuille 2 avec des données de test
        self.timetable.sheet2.append(['Test', 'Test'] + ['X'] * self.timetable.sheet2.max_column)

    def test_create_timetable_automatic(self):
        self.set_up()
        # Appeler la méthode avec l'étudiant de test
        self.timetable.create_timetable_automatic('Test', 'Test', './uploads/test_output.xlsx')

        # Charger la feuille de calcul d'origine et la feuille générée par le test
        original_workbook = openpyxl.load_workbook('./uploads/edt Do_It.23-24.xlsx')
        generated_workbook = openpyxl.load_workbook('./uploads/test_output.xlsx')

        # Vérifier que le fichier a été créé
        self.assertTrue(generated_workbook)

        # Vérifier que chaque cours présent dans la feuille générée était sélectionné par l'étudiant Test Test
        for row in range(1, original_workbook.active.max_row + 1):
            for col in range(1, original_workbook.active.max_column + 1):
                original_cell = original_workbook.active.cell(row=row, column=col).value
                generated_cell = generated_workbook.active.cell(row=row, column=col).value
                # print(f'Comparing cell at ({row}, {col})')
                # print(original_cell)
                if original_cell is not None:
                    msg = (f'\n\nCell at ({row}, {col}) does not match.\n\n'
                           f'Original:\n{original_cell}\n\n'
                           f'Generated:\n{generated_cell}\n')
                    self.assertEqual(original_cell, generated_cell, msg)

    def test_personalize_timetable(self):
        self.timetable = PersonalizeTimetable(
            edt_file_name='./uploads/edt Do_It.23-24.xlsx',
            edt_sheet_name='année',
            students_file_name='./uploads/choix.xlsx',
            students_sheet_name='effectif'
        )
        # Initialiser la feuille 2 avec des données de test
        self.timetable.choices_sheet.append(['Test', 'Empty'] + [''] * self.timetable.choices_sheet.max_column)
        self.timetable.choices_sheet.append(['Test', 'Full'] + ['X'] * self.timetable.choices_sheet.max_column)

        # Appeler la méthode avec l'étudiant de test et tester si le fichier est créé
        self.timetable.create_timetable_automatic('Test', 'Full', './uploads/test_output_full.xlsx')
        self.timetable.create_timetable_automatic('Test', 'Empty', './uploads/test_output_empty.xlsx')
        self.assertTrue(os.path.exists('./uploads/test_output_full.xlsx'))
        self.assertTrue(os.path.exists('./uploads/test_output_empty.xlsx'))

        # Appeler la méthode avec une erreur et assert si l'erreur est relevée
        with self.assertRaises(Exception):
            self.timetable.create_timetable_automatic('unnomparfaitementauhasard',
                                                      'unprenomauhasardaussi',
                                                      './uploads/test_output2.xlsx')

        # Charger la feuille de calcul d'origine et la feuille générée par le test
        original_workbook = openpyxl.load_workbook('./uploads/edt Do_It.23-24.xlsx')
        generated_workbook = openpyxl.load_workbook('./uploads/test_output_full.xlsx')

        # Vérifier que le fichier a été créé
        self.assertTrue(generated_workbook)

        # Vérifier que chaque cours présent dans la feuille générée était sélectionné par l'étudiant Test Test
        for row in range(1, original_workbook.active.max_row + 1):
            for col in range(1, original_workbook.active.max_column + 1):
                original_cell = original_workbook.active.cell(row=row, column=col).value
                generated_cell = generated_workbook.active.cell(row=row, column=col).value
                # print(f'Comparing cell at ({row}, {col})')
                # print(original_cell)
                if original_cell is not None:
                    msg = (f'\n\nCell at ({row}, {col}) does not match.\n\n'
                           f'Original:\n{original_cell}\n\n'
                           f'Generated:\n{generated_cell}\n')
                    self.assertEqual(original_cell, generated_cell, msg)

    def clean_env(self):
        # Supprimer les fichiers de test après chaque test
        files_to_remove = ['test_edt.xlsx', 'test_etudiants.xlsx', 'test_output.xlsx']
        for file_name in files_to_remove:
            try:
                os.remove(file_name)
            except FileNotFoundError:
                pass


if __name__ == '__main__':
    unittest.main()
