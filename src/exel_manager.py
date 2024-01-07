import pandas as pd
import openpyxl
from openpyxl.cell import MergedCell
import requests

from src.data_manager import parse_merged_cell


class ExcelManager:
    def __init__(self, path, sheet_name=""):
        self.sheet_name = sheet_name
        self.xlsx_path = path
        self.workbook = self.load_workbook()

    def load_workbook(self):
        try:
            workbook = openpyxl.load_workbook(self.xlsx_path)
            return workbook
        except Exception as e:
            # Gérez les erreurs liées au chargement du classeur Excel
            print(f"Erreur lors du chargement du classeur Excel : {e}")
            return None

    def excel_to_dataframe(self, sheet_name, start_row=1):
        try:
            sheet = self.workbook[sheet_name]
        except KeyError:
            raise KeyError(f"Sheet '{sheet_name}' not found in the workbook.")

        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            parsed_row = [parse_merged_cell(sheet, row_idx, col_idx + 1) for col_idx, value in enumerate(row)]
            data.append(parsed_row)

        return pd.DataFrame(data)
