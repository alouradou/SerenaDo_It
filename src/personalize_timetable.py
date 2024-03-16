# Copyright Duc Dang Vu
# Original code from
# https://github.com/FrancoisBrucker/do-it/blob/main/src/promos/2023-2024/Dang-Vu-Duc/mon/temps-1.2/create_timetable.py

import openpyxl
import sqlite3


def fetch_courses_aliases():
    conn = sqlite3.connect('../uploads/serenadoit.db')
    cursor = conn.cursor()

    courses_aliases = {}

    # Récupération des données de la base de données
    try:
        cursor.execute("SELECT course_name, alias FROM courses")
        rows = cursor.fetchall()
        for row in rows:
            category, alias = row
            if category in courses_aliases:
                courses_aliases[category].append(alias)
            else:
                courses_aliases[category] = [alias]
    except sqlite3.Error as e:
        print("Erreur lors de la récupération des données depuis la base de données :", e)
    finally:
        conn.close()

    return courses_aliases


def detect_matching(cell, deleted_courses):
    courses_aliases = fetch_courses_aliases()
    flat_list = [alias.lower() for key in deleted_courses for alias in courses_aliases.get(key, [])]
    if cell.value is None:
        return False
    if cell.value.split('\n')[0].lower().strip() in flat_list:
        return True
    return False


def checked_x(value):
    if value is None:
        return False
    if value.lower() == "x":
        return True
    return False


class PersonalizeTimetable:
    def __init__(self,
                 edt_file_name='edt.xlsx', edt_sheet_name='année',
                 students_file_name='etudiants.xlsx', students_sheet_name='effectif'):
        self.edt_file_name = edt_file_name
        self.edt_sheet_name = edt_sheet_name
        self.students_file_name = students_file_name
        self.students_sheet_name = students_sheet_name

        # Initialize workbooks
        self.choices_wb = openpyxl.load_workbook(self.students_file_name)
        self.choices_sheet = self.choices_wb[self.students_sheet_name]
        self.timetable_wb = openpyxl.load_workbook(self.edt_file_name)
        self.timetable_sheet = self.timetable_wb[self.edt_sheet_name]

    def create_timetable_automatic(self, nom, prenom, filename):

        row_number = self.find_student_row(nom, prenom)

        custom_wb = openpyxl.load_workbook(self.edt_file_name)
        custom_sheet = custom_wb[self.edt_sheet_name]

        deleted_courses = []

        if row_number is None:
            raise Exception("Nom ou prénom introuvables")

        # Parcourir toute la feuille de choix pour l'étudiant choisi
        for k in range(7, self.choices_sheet.max_column + 1):
            cell = self.choices_sheet.cell(row=row_number, column=k)
            course_name = str(self.choices_sheet.cell(row=2, column=k).value)
            if course_name != "None" and not checked_x(cell.value):
                # print("Suppression du cours " + course_name)
                deleted_courses.append(course_name.strip())

        # Parcourir tout l'emploi du temps
        for nb_row in range(3, custom_sheet.max_row + 1):
            for nb_column in range(5, custom_sheet.max_column + 1):
                cell = self.timetable_sheet.cell(row=nb_row, column=nb_column)
                if detect_matching(cell, deleted_courses):
                    custom_sheet.cell(row=nb_row, column=nb_column).value = None  # "DELETED " + cell.value
        custom_wb.save(f'{filename}')
        print(f'Emploi du temps enregistré sous le nom de "{filename}"')

    def find_student_row(self, nom, prenom):
        for k in range(3, self.choices_sheet.max_row + 1):
            cell_nom = self.choices_sheet.cell(row=k, column=1)
            cell_prenom = self.choices_sheet.cell(row=k, column=2)

            # Ignorer les lignes vides
            if cell_nom.value is None or cell_prenom.value is None:
                continue

            if prenom.lower() in cell_prenom.value.lower() and nom.lower() in cell_nom.value.lower():
                return k
        print("Nom ou prénom introuvables : " + prenom + " " + nom)
        return None

    def create_all_timetable(self, path="./"):
        for k in range(3, self.choices_sheet.max_row + 1):
            self.choices_wb = openpyxl.load_workbook(self.students_file_name)
            self.choices_sheet = self.choices_wb[self.students_sheet_name]
            self.timetable_wb = openpyxl.load_workbook(self.edt_file_name)
            self.timetable_sheet = self.timetable_wb[self.edt_sheet_name]
            nom = self.choices_sheet.cell(row=k, column=1).value
            prenom = self.choices_sheet.cell(row=k, column=2).value
            self.create_timetable_automatic(nom, prenom, f"{path}edt de {prenom} {nom}.xlsx")
        print("Tout les edt ont été enregistrés")
