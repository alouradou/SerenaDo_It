import os

import pandas as pd
import openpyxl
from openpyxl.cell import MergedCell
import requests
from sqlalchemy import text, create_engine
from datetime import datetime, timedelta


def parse_merged_cell(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # return the left top cell
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value


def compute_timetable_header(df):
    two_first = df.iloc[0:2]
    two_first = two_first.transpose()
    first = two_first[0] + "," + two_first[1]

    df.columns = first
    return df


class DataManager:
    def __init__(self, sheet_id, sheet_name="", saved_workbook_path="./tmp_workbook.xlsx"):
        self.last_fetch = None
        self.db_engine = create_engine("sqlite:///uploads/serenadoit.db")
        self.sheet_id = sheet_id
        self.sheet_name = sheet_name
        self.saved_workbook_path = saved_workbook_path
        # https://stackoverflow.com/questions/33713084/download-link-for-google-spreadsheets-csv-export-with-multiple-sheets
        self.csv_url = "https://docs.google.com/spreadsheets/d/" + sheet_id + "/gviz/tq?tqx=out:csv&sheet=" + sheet_name
        self.xlsx_url = "https://docs.google.com/spreadsheets/d/" + sheet_id + "/export?format=xlsx&id=" + sheet_id
        self.workbook = self.load_workbook()

    # Check if the df is in the db
    def is_sheet_in_cache(self, delta=timedelta(hours=1)):
        try:
            with self.db_engine.connect() as conn:
                fetch_dates = conn.execute(text("SELECT DISTINCT fetch_date "
                                                "FROM fetch_dates "
                                                "WHERE sheet_id = :sheet_id "
                                                "ORDER BY fetch_date"),
                                           {
                                           "sheet_id": self.sheet_id,
                                           "sheet_name": self.sheet_name
                                           }).fetchall()
                self.last_fetch = fetch_dates[-1][0]
                self.last_fetch = datetime.strptime(self.last_fetch, "%Y-%m-%d %H:%M:%S.%f")
                return self.last_fetch > datetime.now() - delta
        except Exception as e:
            print(f"Error checking cache: {e}")
            return False

    def load_workbook(self):
        if self.is_sheet_in_cache():
            print("Loading workbook from cache")
            return self.load_workbook_from_file()
        else:
            print("Loading workbook from internet")
            return self.load_workbook_from_internet()

    def load_workbook_from_file(self):
        try:
            return openpyxl.load_workbook(self.saved_workbook_path, data_only=True)
        except Exception as e:
            raise Exception(f"Error loading workbook from {self.saved_workbook_path}: {e}")

    def load_workbook_from_internet(self):
        file = requests.get(self.xlsx_url, allow_redirects=True)
        # Save date to database table "fetch_dates"
        try:
            df = pd.DataFrame({"fetch_date": [datetime.now()],
                               "sheet_id": [self.sheet_id],
                               "sheet_name": [self.sheet_name]})
            df.to_sql("fetch_dates", self.db_engine, if_exists="append", index=False)
        except Exception as e:
            print(f"Error saving fetch date to database: {e}")
        with open(self.saved_workbook_path, 'wb') as f:
            try:
                f.write(file.content)
                return openpyxl.load_workbook(self.saved_workbook_path, data_only=True)
            except Exception as e:
                raise Exception(f"Error loading workbook from {self.xlsx_url}: {e}")

    def get_from_csv(self):
        return pd.read_csv(self.csv_url)

    def excel_to_dataframe(self, sheet_name, start_row=1):
        try:
            sheet = self.workbook[sheet_name]
        except KeyError:
            raise KeyError(f"Sheet '{sheet_name}' not found in the workbook.")

        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            parsed_row = [parse_merged_cell(sheet, row_idx, col_idx + 1) for col_idx, value in enumerate(row)]
            data.append(parsed_row)

        return pd.DataFrame(data)

# Exemple d'utilisation
# data_manager = DataManager('id_du_fichier')
# event_data = data_manager.get_event_data(0)


class StudentDataManager(DataManager):
    def __init__(self, sheet_id, sheet_name="", saved_workbook_path="./tmp_workbook.xlsx",
                 header_excel_filename="./header_cours_23_24.xlsx"):
        super().__init__(sheet_id, sheet_name, saved_workbook_path)
        self.header_excel_filename = header_excel_filename  # Header contenant le nom des matières
        try:
            self.workbook = self.replace_header(self.workbook, self.sheet_name)
        except FileNotFoundError:
            print("StudentDataManager: Header file not found. Skipping header replacement.")
        print(f"Saving workbook to {saved_workbook_path}")
        self.workbook.save(saved_workbook_path)

    def replace_header(self, workbook, sheet_name):
        # Charger le document Excel avec les deux premières lignes fixes
        header_wb = openpyxl.load_workbook(self.header_excel_filename)
        header_sheet = header_wb.active

        # Remplacer la deuxième ligne par les matières du fichier
        for col in range(1, workbook[sheet_name].max_column + 1):
            workbook[sheet_name].cell(row=2, column=col).value = header_sheet.cell(row=2, column=col).value
            print(f"Replacing column {col} with {header_sheet.cell(row=2, column=col).value}")

        return workbook

    def excel_to_dataframe(self, sheet_name, start_row=1):
        df = super().excel_to_dataframe(sheet_name, start_row)

        # Lower all column names to avoid errors on course name comparison
        df.columns = df.iloc[1]
        df.columns = [col.lower() for col in df.columns]

        try:
            df = df.drop([0, 1]).dropna(subset=["nom", "prénom"])
        except:
            print(df.columns)
            print(df)
            # df = df.drop([0, 1]).dropna(subset=["Nom de famille", "Prénom"])

        return df



